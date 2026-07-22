from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import duckdb
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.ai import embed_texts, llm_json, llm_text
from app.config import get_settings
from app.db import SessionLocal
from app.models import Dataset, DatasetEmbedding, SemanticCatalog


settings = get_settings()


def slugify(value: str, fallback: str = "field") -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_]+", "_", str(value).strip().lower()).strip("_")
    cleaned = re.sub(r"_+", "_", cleaned)
    return cleaned or fallback


def unique_names(values: list[Any], prefix: str = "col") -> list[str]:
    used: dict[str, int] = {}
    result: list[str] = []
    for index, value in enumerate(values, start=1):
        base = slugify(str(value) if value not in (None, "") else "", f"{prefix}_{index}")
        count = used.get(base, 0) + 1
        used[base] = count
        result.append(base if count == 1 else f"{base}_{count}")
    return result


def serialize_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.bool_):
        return bool(value)
    return value


def update_progress(db: Session, dataset_id: str, step: str, detail: str, current: int, total: int = 100) -> None:
    dataset = db.get(Dataset, dataset_id)
    if not dataset:
        return
    dataset.status = "processing"
    dataset.progress_step = step
    dataset.progress_detail = detail
    dataset.progress_current = current
    dataset.progress_total = total
    db.commit()


def update_progress_pct(db: Session, dataset_id: str, step: str, detail: str, pct: float) -> None:
    clamped = max(0, min(100, int(round(pct))))
    update_progress(db, dataset_id, step, detail, clamped, 100)


def infer_semantic_type(column_name: str, series: pd.Series, data_type: str) -> str:
    name = column_name.lower()
    if any(token in name for token in ["fecha", "date", "periodo", "mes", "año", "anio", "year", "month"]):
        return "date"
    if any(token in name for token in ["monto", "venta", "importe", "precio", "costo", "total", "ticket", "margen", "revenue"]):
        return "money_or_amount"
    if any(token in name for token in ["cantidad", "orden", "qty", "unidades", "count"]):
        return "quantity"
    if any(token in name for token in ["id", "codigo", "código", "sku"]):
        return "identifier"
    if "datetime" in data_type or data_type == "date":
        return "date"
    if pd.api.types.is_numeric_dtype(series):
        return "metric"
    non_null = series.dropna()
    if len(non_null) and non_null.nunique(dropna=True) <= max(20, len(non_null) * 0.1):
        return "category"
    return "text"


def profile_column(series: pd.Series) -> dict:
    non_null = series.dropna()
    sample_values = [serialize_value(value) for value in non_null.head(6).tolist()]
    metrics: dict[str, Any] = {
        "nulls": int(series.isna().sum()),
        "non_nulls": int(non_null.shape[0]),
        "distinct": int(non_null.nunique(dropna=True)) if non_null.shape[0] else 0,
    }
    if pd.api.types.is_numeric_dtype(series) and non_null.shape[0]:
        metrics.update(
            {
                "min": serialize_value(non_null.min()),
                "max": serialize_value(non_null.max()),
                "mean": float(non_null.mean()),
            }
        )
    return {"sample_values": sample_values, "metrics": metrics}


def table_name_for_sheet(sheet_name: str) -> str:
    return f"sheet_{slugify(sheet_name, 'sheet')}"


def build_profile(workbook_payload: list[dict], input_kind: str, filename: str) -> dict:
    return {
        "input_kind": input_kind,
        "filename": filename,
        "sheets": [
            {
                "name": item["sheet_name"],
                "duckdb_table": item["duckdb_table"],
                "hidden": item["is_hidden"],
                "rows": item["rows"],
                "columns": item["columns"],
            }
            for item in workbook_payload
        ],
    }


def summarize_metric_value(value: Any, monetary: bool = False) -> str:
    if value is None:
        return "—"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    abs_number = abs(number)
    prefix = "$" if monetary else ""
    if abs_number >= 1_000_000:
        return f"{prefix}{number / 1_000_000:.1f}M"
    if abs_number >= 1_000:
        return f"{prefix}{number / 1_000:.1f}K"
    if number.is_integer():
        return f"{prefix}{int(number):,}"
    return f"{prefix}{number:,.2f}"


def coerce_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    clean = df.copy()
    clean.columns = unique_names([str(col).strip() for col in clean.columns], "col")
    clean = clean.dropna(how="all")
    for column in clean.columns:
        if clean[column].dtype == object:
            converted = pd.to_numeric(clean[column], errors="ignore")
            clean[column] = converted
    return clean


def dataframe_from_sheet(ws) -> tuple[pd.DataFrame, list[str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame(), []
    headers = list(rows[0])
    duck_columns = unique_names(headers)
    original_columns = [
        str(value).strip() if value not in (None, "") else duck_columns[index]
        for index, value in enumerate(headers)
    ]
    df = pd.DataFrame(rows[1:], columns=duck_columns)
    df = coerce_dataframe(df)
    return df, original_columns


def load_csv_payload(file_path: Path) -> list[dict]:
    df = pd.read_csv(file_path, sep=None, engine="python")
    df = coerce_dataframe(df)
    return [
        {
            "sheet_name": file_path.stem,
            "is_hidden": False,
            "columns": list(df.columns),
            "duck_columns": list(df.columns),
            "rows": int(df.shape[0]),
            "df": df,
        }
    ]


def load_excel_payload(file_path: Path) -> list[dict]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    payload: list[dict] = []
    for ws in workbook.worksheets:
        df, original_columns = dataframe_from_sheet(ws)
        payload.append(
            {
                "sheet_name": ws.title,
                "is_hidden": ws.sheet_state != "visible",
                "columns": original_columns,
                "duck_columns": list(df.columns),
                "rows": int(df.shape[0]),
                "df": df,
            }
        )
    return payload


def load_input_payload(file_path: Path, filename: str) -> tuple[str, list[dict]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return "csv", load_csv_payload(file_path)
    if suffix == ".xlsx":
        return "excel", load_excel_payload(file_path)
    raise ValueError("Solo se admiten archivos CSV o Excel (.xlsx).")


def executive_overview_from_payload(workbook_payload: list[dict]) -> list[dict]:
    overview: list[dict] = []
    for sheet in workbook_payload[:4]:
        df = sheet["df"]
        if df.empty:
            continue

        metrics: list[dict[str, Any]] = []
        for column in df.columns:
            series = df[column]
            if not pd.api.types.is_numeric_dtype(series):
                continue
            non_null = series.dropna()
            if non_null.empty:
                continue
            semantic_type = infer_semantic_type(column, series, str(series.dtype))
            metrics.append(
                {
                    "column": column,
                    "label": column.replace("_", " ").title(),
                    "total": float(non_null.sum()),
                    "mean": float(non_null.mean()),
                    "semantic_type": semantic_type,
                }
            )

        dimensions: list[dict[str, Any]] = []
        for column in df.columns:
            if pd.api.types.is_numeric_dtype(df[column]):
                continue
            non_null = df[column].dropna()
            if non_null.empty:
                continue
            distinct = int(non_null.nunique(dropna=True))
            if distinct <= 1:
                continue
            dimensions.append({"column": column, "label": column.replace("_", " ").title(), "distinct": distinct})

        rankings: list[dict[str, Any]] = []
        base_metric = next((item for item in metrics if item["semantic_type"] == "money_or_amount"), metrics[0] if metrics else None)
        for dimension in dimensions[:2]:
            if base_metric:
                grouped = (
                    df[[dimension["column"], base_metric["column"]]]
                    .dropna(subset=[dimension["column"]])
                    .groupby(dimension["column"], dropna=True)[base_metric["column"]]
                    .sum()
                    .sort_values(ascending=False)
                    .head(5)
                )
                rows = [{"label": str(index), "value": float(value)} for index, value in grouped.items()]
            else:
                grouped = df[dimension["column"]].dropna().astype(str).value_counts().head(5)
                rows = [{"label": str(index), "value": int(value)} for index, value in grouped.items()]
            if rows:
                rankings.append(
                    {
                        "dimension": dimension["label"],
                        "metric": base_metric["label"] if base_metric else "Registros",
                        "kind": "bar",
                        "data": rows,
                    }
                )

        overview.append(
            {
                "sheet_name": sheet["sheet_name"],
                "rows": sheet["rows"],
                "metrics": metrics[:4],
                "dimensions": dimensions[:3],
                "rankings": rankings[:2],
            }
        )
    return overview


def fallback_summary(description: str, profile: dict) -> str:
    if description:
        return description
    input_kind = profile.get("input_kind", "archivo")
    total_rows = sum(int(sheet.get("rows") or 0) for sheet in profile.get("sheets", []))
    return f"{input_kind.upper()} listo para análisis con {total_rows:,} registros."


def maybe_llm_summary(description: str, profile: dict) -> str:
    try:
        return llm_text(
            "Resume este archivo para un gerente. Máximo 2 frases. No menciones detalles técnicos.",
            f"Descripción: {description}\nPerfil: {profile}",
        )[:1000]
    except Exception:
        return fallback_summary(description, profile)


def fallback_analysis_json(filename: str, input_kind: str, summary: str, overview: list[dict]) -> dict:
    metrics: list[dict[str, Any]] = []
    insights: list[str] = []
    charts: list[dict[str, Any]] = []

    for sheet in overview[:2]:
        preferred_metrics = [
            metric
            for metric in sheet.get("metrics", [])
            if metric.get("semantic_type") in {"money_or_amount", "metric"}
        ]
        fallback_metrics = [
            metric
            for metric in sheet.get("metrics", [])
            if metric.get("semantic_type") not in {"quantity"}
        ]
        chosen_metrics = (preferred_metrics or fallback_metrics)[:2]
        for metric in chosen_metrics:
            metrics.append(
                {
                    "label": metric["label"],
                    "value": summarize_metric_value(metric["total"], metric["semantic_type"] == "money_or_amount"),
                    "context": "",
                }
            )
        for ranking in sheet.get("rankings", [])[:1]:
            total = sum(float(item["value"]) for item in ranking["data"]) or 0
            if total > 0 and len(ranking["data"]) >= 2:
                for item in ranking["data"][:3]:
                    share = (float(item["value"]) / total) * 100
                    insights.append(f"{item['label']}: {share:.2f}% del total de {ranking['metric']}.")
            else:
                first = ranking["data"][0]
                insights.append(
                    f"En {sheet['sheet_name']}, {first['label']} destaca en {ranking['metric']} con {summarize_metric_value(first['value'])}."
                )
            charts.append(
                {
                    "title": f"{ranking['dimension']} por {ranking['metric']}",
                    "kind": "bar",
                    "data": ranking["data"],
                }
            )

    if not insights:
        insights.append("El archivo ya fue procesado y está listo para análisis conversacional.")

    return {
        "title": filename,
        "input_kind": input_kind,
        "summary": summary,
        "insights": insights[:4],
        "metrics": metrics[:4],
        "charts": charts[:2],
    }


def maybe_analysis_json(filename: str, input_kind: str, summary: str, overview: list[dict]) -> dict:
    fallback = fallback_analysis_json(filename, input_kind, summary, overview)
    try:
        result = llm_json(
            """
Genera solo JSON válido en español para un panel lateral ejecutivo.
Estructura:
{
  "title": string,
  "input_kind": string,
  "summary": string,
  "insights": [string],
  "metrics": [{"label": string, "value": string, "context": string}],
  "charts": [{"title": string, "kind": "bar"|"line", "data": [{"label": string, "value": number}]}]
}
No inventes cifras.
No uses más de 4 métricas ni más de 2 gráficos.
Los KPIs deben ser gerenciales. Prioriza montos, totales y promedios relevantes.
No uses conteos de filas o registros como KPI principal salvo que no exista otra métrica útil.
Evita textos como "Hoja1" o nombres técnicos en el contexto del KPI. Si no aporta valor, deja context vacío.
Como el panel lateral es compacto, prioriza gráficos "bar" horizontales. Evita "pie".
Cuando una distribución tenga 2 o más categorías relevantes, expresa los insights como puntos cortos, uno por categoría principal.
""",
            f"Archivo: {filename}\nTipo: {input_kind}\nResumen: {summary}\nOverview: {json.dumps(overview, ensure_ascii=False)}",
        )
        result.setdefault("title", filename)
        result.setdefault("input_kind", input_kind)
        result.setdefault("summary", summary)
        result["insights"] = [str(item) for item in (result.get("insights") or [])][:4]
        result["metrics"] = list(result.get("metrics") or [])[:4]
        result["charts"] = list(result.get("charts") or [])[:2]
        for metric in result["metrics"]:
            if any(token in str(metric.get("context") or "").lower() for token in ["hoja", "sheet"]):
                metric["context"] = ""
        return result
    except Exception:
        return fallback


def chunk_text(text: str, max_chars: int = 12000) -> str:
    return text if len(text) <= max_chars else text[:max_chars] + "\n[truncado para embedding]"


def build_embedding_chunks(description: str, profile: dict, workbook_payload: list[dict], analysis_json: dict) -> list[dict]:
    chunks: list[dict] = [
        {
            "chunk_type": "workbook_summary",
            "sheet_name": None,
            "chunk_index": 0,
            "content": (
                f"Descripción del usuario:\n{description}\n\n"
                f"Perfil del archivo:\n{profile}\n\n"
                f"Análisis estructurado:\n{json.dumps(analysis_json, ensure_ascii=False)}"
            ),
            "metadata_json": {"profile": True},
        }
    ]
    for sheet in workbook_payload:
        chunks.append(
            {
                "chunk_type": "sheet_catalog",
                "sheet_name": sheet["sheet_name"],
                "chunk_index": 0,
                "content": (
                    f"Sección: {sheet['sheet_name']}\n"
                    f"Tabla DuckDB: {sheet['duckdb_table']}\n"
                    f"Columnas: {sheet['columns']}"
                ),
                "metadata_json": {"duckdb_table": sheet["duckdb_table"]},
            }
        )
        df = sheet["df"]
        for start in range(0, len(df), settings.row_chunk_size):
            chunk_df = df.iloc[start : start + settings.row_chunk_size]
            content = (
                f"Sección: {sheet['sheet_name']}\n"
                f"Tabla DuckDB: {sheet['duckdb_table']}\n"
                f"Filas aproximadas: {start + 1} a {start + len(chunk_df)}\n\n"
                f"{chunk_df.to_csv(index=False)}"
            )
            chunks.append(
                {
                    "chunk_type": "row_chunk",
                    "sheet_name": sheet["sheet_name"],
                    "chunk_index": start // settings.row_chunk_size,
                    "content": chunk_text(content),
                    "metadata_json": {
                        "duckdb_table": sheet["duckdb_table"],
                        "start_row": start + 1,
                        "end_row": start + len(chunk_df),
                    },
                }
            )
    return chunks


def save_embeddings(db: Session, dataset_id: str, chunks: list[dict]) -> None:
    db.query(DatasetEmbedding).filter(DatasetEmbedding.dataset_id == dataset_id).delete()
    batch_size = 32
    total_batches = max(1, (len(chunks) + batch_size - 1) // batch_size)
    for start in range(0, len(chunks), batch_size):
        batch_number = start // batch_size + 1
        pct = 72 + ((batch_number / total_batches) * 24)
        update_progress_pct(
            db,
            dataset_id,
            "embeddings",
            f"Generando índice semántico ({batch_number}/{total_batches}).",
            pct,
        )
        batch = chunks[start : start + batch_size]
        vectors = embed_texts([item["content"] for item in batch])
        for item, vector in zip(batch, vectors):
            db.add(
                DatasetEmbedding(
                    dataset_id=dataset_id,
                    chunk_type=item["chunk_type"],
                    sheet_name=item["sheet_name"],
                    chunk_index=item["chunk_index"],
                    content=item["content"],
                    metadata_json=item["metadata_json"],
                    embedding=vector,
                )
            )
        db.commit()


def process_dataset(dataset_id: str, file_path: Path) -> None:
    db = SessionLocal()
    con = None
    try:
        dataset = db.get(Dataset, dataset_id)
        if not dataset:
            return

        dataset.error_message = None
        dataset.summary = None
        dataset.profile_json = None
        db.commit()

        update_progress_pct(db, dataset_id, "reading_file", "Leyendo el archivo cargado.", 5)
        input_kind, raw_payload = load_input_payload(file_path, dataset.filename)

        duckdb_path = settings.storage_dir / "duckdb" / f"{dataset_id}.duckdb"
        if duckdb_path.exists():
            duckdb_path.unlink()
        con = duckdb.connect(str(duckdb_path))

        update_progress_pct(db, dataset_id, "catalog", "Creando catálogo semántico.", 12)
        db.query(SemanticCatalog).filter(SemanticCatalog.dataset_id == dataset_id).delete()
        db.query(DatasetEmbedding).filter(DatasetEmbedding.dataset_id == dataset_id).delete()
        db.commit()

        workbook_payload: list[dict] = []
        total_sections = len(raw_payload)
        for section_index, section in enumerate(raw_payload, start=1):
            section_pct = 18 + ((section_index - 1) / max(1, total_sections)) * 36
            update_progress_pct(
                db,
                dataset_id,
                "duckdb",
                f"Preparando sección {section_index}/{total_sections}: {section['sheet_name']}.",
                section_pct,
            )
            df = section["df"]
            duckdb_table = table_name_for_sheet(section["sheet_name"])
            con.register("input_df", df)
            con.execute(f'CREATE TABLE "{duckdb_table}" AS SELECT * FROM input_df')
            con.unregister("input_df")
            workbook_payload.append(
                {
                    "sheet_name": section["sheet_name"],
                    "duckdb_table": duckdb_table,
                    "is_hidden": section["is_hidden"],
                    "rows": section["rows"],
                    "columns": section["columns"],
                    "duck_columns": section["duck_columns"],
                    "df": df,
                }
            )

            for original_name, duck_name in zip(section["columns"], section["duck_columns"]):
                series = df[duck_name]
                data_type = str(series.infer_objects().dtype)
                col_profile = profile_column(series)
                semantic_type = infer_semantic_type(original_name, series, data_type)
                db.add(
                    SemanticCatalog(
                        dataset_id=dataset_id,
                        sheet_name=section["sheet_name"],
                        duckdb_table=duckdb_table,
                        column_name=original_name,
                        duckdb_column=duck_name,
                        data_type=data_type,
                        semantic_type=semantic_type,
                        description=f"Campo {original_name} detectado como {semantic_type}.",
                        sample_values=col_profile["sample_values"],
                        metrics_json=col_profile["metrics"],
                        is_hidden=section["is_hidden"],
                    )
                )
            db.commit()
            finished_section_pct = 18 + (section_index / max(1, total_sections)) * 36
            update_progress_pct(
                db,
                dataset_id,
                "duckdb",
                f"Sección {section_index}/{total_sections} lista: {section['sheet_name']}.",
                finished_section_pct,
            )

        if con:
            con.close()
            con = None

        update_progress_pct(db, dataset_id, "summary", "Generando análisis estructurado.", 60)
        profile = build_profile(workbook_payload, input_kind, dataset.filename)
        overview = executive_overview_from_payload(workbook_payload)
        dataset = db.get(Dataset, dataset_id)
        dataset.summary = maybe_llm_summary(dataset.description or dataset.filename, profile)
        analysis_json = maybe_analysis_json(dataset.filename, input_kind, dataset.summary or "", overview)
        dataset.profile_json = {
            **profile,
            "executive_overview": overview,
            "analysis_json": analysis_json,
        }
        db.commit()

        update_progress_pct(db, dataset_id, "chunking", "Preparando contexto para preguntas.", 70)
        chunks = build_embedding_chunks(dataset.description or "", profile, workbook_payload, analysis_json)
        save_embeddings(db, dataset_id, chunks)

        dataset = db.get(Dataset, dataset_id)
        dataset.status = "ready"
        dataset.progress_step = "ready"
        dataset.progress_detail = "Archivo listo para conversar."
        dataset.progress_current = 100
        dataset.progress_total = 100
        db.commit()
    except Exception as exc:
        db.rollback()
        dataset = db.get(Dataset, dataset_id)
        if dataset:
            dataset.status = "error"
            dataset.error_message = str(exc)
            dataset.progress_step = "error"
            dataset.progress_detail = str(exc)
            db.commit()
    finally:
        if con:
            con.close()
        db.close()
